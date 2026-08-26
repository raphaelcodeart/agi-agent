"""Piccolo helper generico per generare un export .xlsx in memoria (nessun
file temporaneo su disco), usato dai 3 bottoni "Esporta Excel" del modulo
Statistiche (vedi app/api/v1/statistics.py e docs/STATISTICS.md)."""
import io
from collections.abc import Sequence
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def build_xlsx(sheet_title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # limite Excel sul nome del foglio

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            # Excel/openpyxl non accetta datetime timezone-aware
            if isinstance(value, datetime) and value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
